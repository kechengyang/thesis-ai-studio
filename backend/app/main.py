from __future__ import annotations

import csv
import json
import os
import re
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from .analysis_skills import (
    chat_brief_turn,
    chat_data_turn,
    chat_mindmap_turn,
    dataframe_profile,
    insert_block_into_section,
    insert_figure_into_manuscript,
    insert_mermaid_into_manuscript,
    load_tabular_data,
    run_brief_skill,
    run_data_analysis_skill,
    run_mindmap_skill,
    safe_project_file,
    SUPPORTED_DATA_SUFFIXES,
)
from .config import (
    DEFAULT_AI_INSTRUCTION,
    DEFAULT_PROJECT_ID,
    DEFAULT_PROJECTS_ROOT,
    DEFAULT_PROVIDER_MODELS,
    ENV_PATH,
    LEGACY_WORKSPACE,
    PROJECT_FOLDERS,
    STATE_PATH,
    load_env_file as config_load_env_file,
    mask_secret,
    read_settings as config_read_settings,
    save_settings as config_save_settings,
    settings_payload as config_settings_payload,
    update_env_values,
)
from .literature import (
    build_google_scholar_search_url,
    build_literature_prompt,
    cache_literature_result,
    import_literature_source,
    load_cached_literature,
    save_literature_review_output,
    search_literature_candidates,
)
from .providers import get_provider, provider_payload
from .schemas import (
    ApplyRequest,
    BriefRequest,
    ChatRequest,
    DataAnalysisRequest,
    DataFigureInsertRequest,
    DocumentCreate,
    DocumentOpen,
    DocumentUpdate,
    LiteratureAnalyzeRequest,
    LiteratureImportRequest,
    MindmapInsertRequest,
    MindmapRequest,
    ProjectCreate,
    ProjectFileDelete,
    ProjectFileMove,
    ProjectFileRename,
    ProjectOpen,
    RejectRequest,
    SettingsUpdate,
    SuggestRequest,
    WorkspaceRootUpdate,
)

DEFAULT_PAPER = """---
title: "Working Paper Title"
author: "Author Name"
format:
  docx:
    toc: true
bibliography: references.bib
---

# Abstract

Write your abstract here. Select any paragraph and ask the AI panel to revise, expand, or check the logic.

# Introduction

Introduce the research background, research question, and contribution.

# Methods

Describe the research design, data sources, and analytical strategy.

# Results

Present the main findings. Figures can be placed in the `figures/` folder and referenced with Quarto syntax.

# Discussion

Interpret the findings, discuss limitations, and connect the argument to the literature.

# References
"""

app = FastAPI(title="Quarto AI Paper Studio", version="0.1.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "null",  # Electron file:// pages send Origin: null
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "http://localhost:5183",
        "http://127.0.0.1:5183",
    ],
    allow_origin_regex=r"http://(localhost|127\.0\.0\.1):\d+",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

WORKSPACE_REQUIRED_MESSAGE = "请先选择 Workspace 文件夹。"
WORKSPACE_MISSING_MESSAGE = "当前 Workspace 文件夹不存在，请重新选择。"
WORKSPACE_INVALID_MESSAGE = "当前 Workspace 路径不是文件夹，请重新选择。"
CATEGORY_DIRECTORY_MAP = {
    "Project Root": "",
    "Data": "data",
    "Figures": "figures",
    "Templates": "templates",
    "Outputs": "outputs",
}
DRAGGABLE_CATEGORIES = tuple(CATEGORY_DIRECTORY_MAP.keys())


def slugify(value: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9._-]+", "-", value.strip().lower()).strip("-._")
    return slug[:80] or DEFAULT_PROJECT_ID


def ensure_settings() -> None:
    config_read_settings()


def read_state() -> dict[str, Any]:
    STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    if not STATE_PATH.exists():
        return {}
    try:
        return json.loads(STATE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def write_state(state: dict[str, Any]) -> None:
    STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    STATE_PATH.write_text(json.dumps(state, indent=2, ensure_ascii=False), encoding="utf-8")


def merge_state(**updates: Any) -> dict[str, Any]:
    state = read_state()
    state.update(updates)
    write_state(state)
    return state


def workspace_state_key(path: Optional[Path] = None) -> str:
    return str((path or workspace_path()).resolve())


def configured_projects_root() -> Optional[Path]:
    env_root = str(os.environ.get("THESIS_PROJECTS_ROOT", "")).strip()
    if env_root:
        return Path(env_root).expanduser().resolve()
    state = read_state()
    remembered_workspace = str(state.get("workspace_path", "")).strip()
    if remembered_workspace:
        return Path(remembered_workspace).expanduser().resolve()
    remembered_root = str(state.get("projects_root", "")).strip()
    if remembered_root:
        return Path(remembered_root).expanduser().resolve()
    return None


def suggested_projects_root() -> Path:
    return (configured_projects_root() or DEFAULT_PROJECTS_ROOT).resolve()


def workspace_label(path: Optional[Path] = None) -> str:
    target = (path or configured_projects_root() or DEFAULT_PROJECTS_ROOT).resolve()
    return target.name or DEFAULT_PROJECT_ID


def projects_root_path() -> Path:
    root = configured_projects_root()
    if root is None:
        raise HTTPException(status_code=409, detail=WORKSPACE_REQUIRED_MESSAGE)
    return root


def normalize_projects_root(root_path: str) -> Path:
    cleaned = str(root_path or "").strip()
    if not cleaned:
        raise HTTPException(status_code=400, detail="请选择一个 Workspace 文件夹。")
    return Path(cleaned).expanduser()


def safe_project_path(project_id: str, root: Optional[Path] = None) -> Path:
    projects_root = (root or projects_root_path()).expanduser()
    project_path = (projects_root / slugify(project_id)).resolve()
    root = projects_root.resolve()
    try:
        project_path.relative_to(root)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="项目名称不安全，请换一个名称。") from exc
    return project_path


def manuscript_paths(project: Path) -> list[Path]:
    return sorted(path for path in project.iterdir() if path.is_file() and path.suffix.lower() == ".qmd")


def manuscript_entries(project_id: Optional[str] = None) -> list[dict[str, Any]]:
    project = workspace_path(project_id)
    return [
        {
            "name": path.name,
            "relative_path": path.name,
            "size": path.stat().st_size,
            "size_label": file_size_label(path.stat().st_size),
            "modified_at": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
        }
        for path in manuscript_paths(project)
    ]


def default_manuscript_name(project: Path) -> str:
    manuscripts = manuscript_paths(project)
    if not manuscripts:
        return "paper.qmd"
    preferred = project / "paper.qmd"
    if preferred.exists():
        return preferred.name
    return manuscripts[0].name


def active_manuscript_relative_path(project_id: Optional[str] = None) -> str:
    project = workspace_path(project_id)
    project_key = workspace_state_key(project)
    state = read_state()
    active_manuscripts = state.get("active_manuscripts", {})
    candidate = str(active_manuscripts.get(project_key, "")).strip()
    if candidate:
        path = project / candidate
        if path.exists() and path.is_file() and path.suffix.lower() == ".qmd":
            return candidate
    fallback = default_manuscript_name(project)
    active_manuscripts[project_key] = fallback
    merge_state(active_manuscripts=active_manuscripts)
    return fallback


def set_active_manuscript(relative_path: str, project_id: Optional[str] = None) -> str:
    project = workspace_path(project_id)
    cleaned = Path(relative_path.strip()).name
    if not cleaned.lower().endswith(".qmd"):
        raise HTTPException(status_code=400, detail="只能切换到 qmd 文稿。")
    target = project / cleaned
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="没有找到这个 qmd 文稿。")
    project_key = workspace_state_key(project)
    state = read_state()
    active_manuscripts = state.get("active_manuscripts", {})
    active_manuscripts[project_key] = cleaned
    merge_state(active_manuscripts=active_manuscripts)
    return cleaned


def paper_path(project_id: Optional[str] = None) -> Path:
    project = workspace_path(project_id)
    return project / active_manuscript_relative_path(project_id)


def create_manuscript_file(filename: str, project_id: Optional[str] = None) -> Path:
    project = workspace_path(project_id)
    raw_name = Path(filename.strip()).stem
    safe_name = slugify(raw_name)
    if not safe_name:
        raise HTTPException(status_code=400, detail="请输入新的 qmd 文件名。")
    target = project / f"{safe_name}.qmd"
    if target.exists():
        raise HTTPException(status_code=409, detail="这个 qmd 文件已经存在。")
    title = safe_name.replace("-", " ").title()
    target.write_text(DEFAULT_PAPER.replace('title: "Working Paper Title"', f'title: "{title}"'), encoding="utf-8")
    set_active_manuscript(target.name, project_id)
    return target


def project_file_path(relative_path: str, project_id: Optional[str] = None) -> Path:
    project = workspace_path(project_id).resolve()
    cleaned = relative_path.strip().lstrip("/")
    if not cleaned:
        raise HTTPException(status_code=400, detail="缺少文件路径。")
    candidate = (project / cleaned).resolve()
    try:
        candidate.relative_to(project)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="文件路径不安全。") from exc
    if not candidate.exists() or not candidate.is_file():
        raise HTTPException(status_code=404, detail="没有找到这个项目文件。")
    return candidate


def source_index_path(project_id: Optional[str] = None) -> Path:
    return workspace_path(project_id) / "sources" / "sources_index.json"


def memory_dir(project_id: Optional[str] = None) -> Path:
    return workspace_path(project_id) / "memory"


def memory_path(name: str, project_id: Optional[str] = None) -> Path:
    return memory_dir(project_id) / name


def existing_projects(root: Optional[Path] = None) -> list[dict[str, Any]]:
    workspace = (root or workspace_path()).expanduser().resolve()
    if not workspace.exists() or not workspace.is_dir():
        return []
    manuscripts = manuscript_paths(workspace)
    modified_source = max(manuscripts, key=lambda item: item.stat().st_mtime) if manuscripts else workspace
    return [
        {
            "id": workspace_label(workspace),
            "name": workspace_label(workspace),
            "path": str(workspace),
            "paper_exists": bool(manuscripts),
            "modified_at": datetime.fromtimestamp(modified_source.stat().st_mtime).isoformat(timespec="seconds"),
        }
    ]


def copy_legacy_materials(destination: Path) -> None:
    if not LEGACY_WORKSPACE.exists():
        return
    for folder in PROJECT_FOLDERS:
        source_dir = LEGACY_WORKSPACE / folder
        target_dir = destination / folder
        if not source_dir.exists() or not source_dir.is_dir():
            continue
        for source_file in source_dir.iterdir():
            if not source_file.is_file() or source_file.name == ".gitkeep":
                continue
            target_file = target_dir / source_file.name
            if not target_file.exists():
                shutil.copy2(source_file, target_file)


def repair_source_index(project: Path) -> None:
    index_path = project / "sources" / "sources_index.json"
    try:
        existing = json.loads(index_path.read_text(encoding="utf-8")) if index_path.exists() else []
    except json.JSONDecodeError:
        existing = []
    if existing:
        return
    entries = []
    candidate_texts = list((project / "sources").glob("*.txt")) if (project / "sources").exists() else []
    data_dir = project / "data"
    if data_dir.exists():
        for data_file in sorted(data_dir.iterdir()):
            if not data_file.is_file() or data_file.name == ".DS_Store":
                continue
            if data_file.suffix.lower() not in [".csv", ".xlsx", ".xlsm"]:
                continue
            text_file = project / "sources" / f"data-{data_file.stem}.txt"
            if not text_file.exists():
                try:
                    text_file.write_text(extract_text(data_file), encoding="utf-8")
                except HTTPException:
                    continue
            candidate_texts.append(text_file)
    if not candidate_texts:
        return
    for text_path in sorted(set(candidate_texts)):
        source_dir = text_path.parent
        originals = [
            path
            for path in source_dir.iterdir()
            if path.is_file()
            and path.stem == text_path.stem
            and path.name != text_path.name
            and path.suffix.lower() in [".pdf", ".docx", ".csv", ".xlsx", ".xlsm"]
        ]
        if text_path.name.startswith("data-"):
            data_name = text_path.name.removeprefix("data-").removesuffix(".txt")
            originals = [
                path
                for path in data_dir.iterdir()
                if path.is_file() and path.stem == data_name and path.suffix.lower() in [".csv", ".xlsx", ".xlsm"]
            ]
        filename = originals[0].name if originals else text_path.name
        text = text_path.read_text(encoding="utf-8", errors="ignore")
        entries.append(
            {
                "filename": filename,
                "text_file": text_path.name,
                "characters": len(text),
                "imported_at": datetime.fromtimestamp(text_path.stat().st_mtime).isoformat(timespec="seconds"),
            }
        )
    if entries:
        index_path.write_text(json.dumps(entries, indent=2, ensure_ascii=False), encoding="utf-8")


def ensure_memory_files(project: Path) -> None:
    memory = project / "memory"
    memory.mkdir(parents=True, exist_ok=True)
    for filename in ["conversations.jsonl", "changes.jsonl"]:
        path = memory / filename
        if not path.exists():
            path.write_text("", encoding="utf-8")
    summary = memory / "summary.md"
    if not summary.exists():
        summary.write_text(
            "# Project Memory\n\n"
            "This file is maintained by the local AI paper studio. It summarizes AI interactions and accepted edits for this project.\n",
            encoding="utf-8",
        )


def ensure_project_scaffold(project: Path) -> None:
    for relative in [
        "data",
        "sources",
        "figures",
        "templates",
        "outputs",
        "outputs/analysis",
        "outputs/briefs",
        "outputs/mindmaps",
        "memory",
        "memory/chats",
    ]:
        (project / relative).mkdir(parents=True, exist_ok=True)


def ensure_project(
    project_id: Optional[str] = None,
    title: Optional[str] = None,
    seed_legacy: bool = False,
    root: Optional[Path] = None,
) -> Path:
    ensure_settings()
    project = (root or projects_root_path()).expanduser().resolve()
    project.mkdir(parents=True, exist_ok=True)
    ensure_project_scaffold(project)
    if not manuscript_paths(project):
        paper = DEFAULT_PAPER
        if title:
            paper = paper.replace('title: "Working Paper Title"', f'title: "{title}"')
        (project / "paper.qmd").write_text(paper, encoding="utf-8")
    if not (project / "references.bib").exists():
        (project / "references.bib").write_text("", encoding="utf-8")
    if not (project / "sources" / "sources_index.json").exists():
        (project / "sources" / "sources_index.json").write_text("[]", encoding="utf-8")
    if seed_legacy:
        copy_legacy_materials(project)
    repair_source_index(project)
    ensure_memory_files(project)
    return project


def resolve_active_project_id(root: Optional[Path] = None, preferred: Optional[str] = None) -> str:
    project = ensure_project(title=preferred or workspace_label(root or projects_root_path()), root=root or projects_root_path())
    return workspace_label(project)


def active_project_id() -> str:
    state = read_state()
    selected = resolve_active_project_id(preferred=state.get("active_project", workspace_label(projects_root_path())))
    updates: dict[str, Any] = {"active_project": selected}
    if "THESIS_PROJECTS_ROOT" not in os.environ:
        updates["workspace_path"] = str(projects_root_path())
    merge_state(**updates)
    return selected


def set_active_project(project_id: str) -> None:
    target = (workspace_path().parent / slugify(project_id)).resolve()
    if not target.exists() or not target.is_dir():
        raise HTTPException(status_code=404, detail="没有找到这个 Workspace 文件夹。")
    set_projects_root(str(target))


def set_projects_root(root_path: str) -> Path:
    if "THESIS_PROJECTS_ROOT" in os.environ:
        raise HTTPException(status_code=409, detail="当前通过 THESIS_PROJECTS_ROOT 启动，界面里不能修改 Workspace 文件夹。")
    project = normalize_projects_root(root_path)
    if project.exists() and not project.is_dir():
        raise HTTPException(status_code=400, detail="所选路径不是文件夹。")
    project.mkdir(parents=True, exist_ok=True)
    project = project.resolve()
    ensure_project(title=workspace_label(project), root=project)
    merge_state(workspace_path=str(project), projects_root=str(project), active_project=workspace_label(project))
    return project


def workspace_path(project_id: Optional[str] = None) -> Path:
    project = projects_root_path()
    if project.exists() and not project.is_dir():
        raise HTTPException(status_code=409, detail=WORKSPACE_INVALID_MESSAGE)
    if not project.exists():
        raise HTTPException(status_code=409, detail=WORKSPACE_MISSING_MESSAGE)
    ensure_project(title=workspace_label(project), root=project)
    return project


def ensure_workspace() -> None:
    project = workspace_path()
    active_project_id()
    if "THESIS_PROJECTS_ROOT" not in os.environ:
        merge_state(workspace_path=str(project), projects_root=str(project))


def load_env_file() -> dict[str, str]:
    return config_load_env_file()


def read_settings() -> dict[str, Any]:
    return config_read_settings()


def save_settings(settings: dict[str, Any]) -> None:
    config_save_settings(settings)


def save_api_key(api_key: str) -> None:
    update_env_values({"OPENAI_API_KEY": api_key})


def masked_api_key(provider: Optional[str] = None) -> str:
    settings = read_settings()
    env = load_env_file()
    if provider == "deepseek":
        return mask_secret(os.environ.get("DEEPSEEK_API_KEY") or env.get("DEEPSEEK_API_KEY", ""))
    if provider == "openai":
        return mask_secret(os.environ.get("OPENAI_API_KEY") or env.get("OPENAI_API_KEY", ""))
    active_provider = provider or settings.get("provider", "openai")
    if active_provider == "deepseek":
        return mask_secret(os.environ.get("DEEPSEEK_API_KEY") or env.get("DEEPSEEK_API_KEY", ""))
    return mask_secret(os.environ.get("OPENAI_API_KEY") or env.get("OPENAI_API_KEY", ""))


def choose_projects_root_dialog() -> Optional[Path]:
    initial_dir = suggested_projects_root()
    if sys.platform == "darwin":
        escaped_initial_dir = str(initial_dir).replace("\\", "\\\\").replace('"', '\\"')
        script_lines = ['set promptText to "Choose a Workspace folder"']
        if initial_dir.exists():
            script_lines.append(f'set defaultFolder to POSIX file "{escaped_initial_dir}"')
            script_lines.append("set chosenFolder to choose folder with prompt promptText default location defaultFolder")
        else:
            script_lines.append("set chosenFolder to choose folder with prompt promptText")
        script_lines.append("POSIX path of chosenFolder")
        result = subprocess.run(
            ["osascript", *sum((["-e", line] for line in script_lines), [])],
            capture_output=True,
            text=True,
            check=False,
        )
        if result.returncode != 0:
            return None
        selected = result.stdout.strip()
        return Path(selected).expanduser() if selected else None
    raise HTTPException(status_code=501, detail="当前系统暂未实现文件夹选择器，请手动输入路径。")


def outline_from_document(content: str) -> list[dict[str, Any]]:
    outline = []
    for line in content.splitlines():
        match = re.match(r"^(#{1,6})\s+(.+)$", line.strip())
        if match:
            outline.append({"level": len(match.group(1)), "title": match.group(2)})
    return outline


def read_source_index(project_id: Optional[str] = None) -> list[dict[str, Any]]:
    ensure_workspace()
    index_path = source_index_path(project_id)
    try:
        return json.loads(index_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return []


def write_source_index(items: list[dict[str, Any]], project_id: Optional[str] = None) -> None:
    source_index_path(project_id).write_text(json.dumps(items, indent=2, ensure_ascii=False), encoding="utf-8")


def file_size_label(size: int) -> str:
    if size < 1024:
        return f"{size} B"
    if size < 1024 * 1024:
        return f"{size / 1024:.1f} KB"
    return f"{size / (1024 * 1024):.1f} MB"


def project_relative_path(path: Path, project: Path) -> str:
    return str(path.relative_to(project))


def list_source_entries(project: Path) -> list[dict[str, Any]]:
    index_path = project / "sources" / "sources_index.json"
    if not index_path.exists():
        return []
    try:
        items = json.loads(index_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return []
    if not isinstance(items, list):
        return []
    return [item for item in items if isinstance(item, dict)]


def write_source_entries(project: Path, items: list[dict[str, Any]]) -> None:
    (project / "sources" / "sources_index.json").write_text(json.dumps(items, indent=2, ensure_ascii=False), encoding="utf-8")


def source_entry_for_path(project: Path, path: Path) -> dict[str, Any] | None:
    name = path.name
    for item in list_source_entries(project):
        if item.get("filename") == name or item.get("text_file") == name:
            return item
    return None


def project_file_category(project: Path, path: Path) -> str:
    relative = path.relative_to(project)
    parts = relative.parts
    if len(parts) == 1:
        if path.suffix.lower() == ".qmd" or path.name == "references.bib":
            return "Manuscript"
        return "Project Root"
    if parts[0] == "sources":
        return "Sources"
    if parts[0] == "data":
        return "Data"
    if parts[0] == "figures":
        return "Figures"
    if parts[0] == "templates":
        return "Templates"
    if parts[0] == "outputs":
        return "Outputs"
    if parts[0] == "memory":
        return "Memory"
    return "Project Root"


def move_targets_for_file(category: str, source_entry: dict[str, Any] | None, path: Path) -> list[str]:
    if category not in DRAGGABLE_CATEGORIES:
        return []
    if source_entry:
        return []
    if path.suffix.lower() == ".qmd":
        return []
    return [item for item in DRAGGABLE_CATEGORIES if item != category]


def project_file_actions(project: Path, path: Path, category: str) -> dict[str, Any]:
    source_entry = source_entry_for_path(project, path) if category == "Sources" else None
    source_filename = str(source_entry.get("filename", "")).strip() if source_entry else ""
    source_text_file = str(source_entry.get("text_file", "")).strip() if source_entry else ""
    is_source_sidecar = bool(source_entry and path.name == source_text_file and source_filename and source_filename != source_text_file)
    is_manuscript = category == "Manuscript" and path.suffix.lower() == ".qmd"
    is_protected = (
        path.name == "references.bib"
        or category == "Memory"
        or is_source_sidecar
    )
    can_delete = not is_protected
    if is_manuscript and len(manuscript_paths(project)) <= 1:
        can_delete = False
    can_rename = not is_protected
    can_move = bool(move_targets_for_file(category, source_entry, path))
    return {
        "can_rename": can_rename,
        "can_delete": can_delete,
        "can_move": can_move,
        "move_targets": move_targets_for_file(category, source_entry, path),
    }


def list_project_files(project: Path) -> list[dict[str, Any]]:
    manuscript_files = manuscript_paths(project)
    manuscript_names = {path.name for path in manuscript_files}
    categories = [
        ("Manuscript", manuscript_files + ([project / "references.bib"] if (project / "references.bib").exists() else [])),
        (
            "Project Root",
            sorted(
                path
                for path in project.iterdir()
                if path.is_file() and path.name not in manuscript_names and path.name != "references.bib"
            ),
        ),
        ("Sources", sorted((project / "sources").glob("*"))),
        ("Data", sorted((project / "data").glob("*"))),
        ("Figures", sorted((project / "figures").glob("*"))),
        ("Templates", sorted((project / "templates").glob("*"))),
        ("Outputs", sorted((project / "outputs").glob("*"))),
        ("Memory", sorted((project / "memory").glob("*"))),
    ]
    result = []
    for category, paths in categories:
        files = []
        for path in paths:
            if not path.exists() or not path.is_file() or path.name == ".gitkeep":
                continue
            if path.name == "sources_index.json":
                continue
            files.append(
                {
                    "name": path.name,
                    "relative_path": str(path.relative_to(project)),
                    "size": path.stat().st_size,
                    "size_label": file_size_label(path.stat().st_size),
                    "modified_at": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
                    "extension": path.suffix.lower().lstrip(".") or "file",
                    **project_file_actions(project, path, category),
                }
            )
        result.append({"category": category, "files": files})
    return result


def ensure_same_project_path(project: Path, path: Path) -> None:
    try:
        path.resolve().relative_to(project.resolve())
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="文件路径不安全。") from exc


def target_directory_for_category(project: Path, category: str) -> Path:
    if category not in CATEGORY_DIRECTORY_MAP:
        raise HTTPException(status_code=400, detail="不能拖到这个分组。")
    relative = CATEGORY_DIRECTORY_MAP[category]
    return project if not relative else project / relative


def requested_filename(path: Path, new_name: str) -> str:
    candidate = Path(str(new_name or "").strip()).name.strip()
    if not candidate or candidate in {".", ".."}:
        raise HTTPException(status_code=400, detail="请输入有效的文件名。")
    if path.suffix.lower() == ".qmd":
        stem = slugify(Path(candidate).stem)
        if not stem:
            raise HTTPException(status_code=400, detail="请输入有效的 qmd 文件名。")
        return f"{stem}.qmd"
    if "." not in Path(candidate).name and path.suffix:
        return f"{candidate}{path.suffix}"
    return candidate


def update_active_manuscript_reference(project: Path, old_name: str, new_name: str | None = None) -> None:
    project_key = workspace_state_key(project)
    state = read_state()
    active_manuscripts = state.get("active_manuscripts", {})
    current = str(active_manuscripts.get(project_key, "")).strip()
    if current != old_name:
        return
    active_manuscripts[project_key] = new_name or default_manuscript_name(project)
    merge_state(active_manuscripts=active_manuscripts)


def migrate_editor_chat_history(project: Path, old_name: str, new_name: str | None = None) -> None:
    old_path = chat_history_path("editor", project, chat_key=old_name)
    if not old_path.exists():
        return
    if not new_name:
        old_path.unlink(missing_ok=True)
        return
    new_path = chat_history_path("editor", project, chat_key=new_name)
    if new_path == old_path:
        return
    new_path.parent.mkdir(parents=True, exist_ok=True)
    if new_path.exists():
        with new_path.open("a", encoding="utf-8") as handle:
            handle.write(old_path.read_text(encoding="utf-8", errors="ignore"))
        old_path.unlink(missing_ok=True)
        return
    old_path.rename(new_path)


def rename_source_managed_file(project: Path, entry: dict[str, Any], new_name: str) -> dict[str, Any]:
    original_name = str(entry.get("filename", "")).strip()
    text_name = str(entry.get("text_file", "")).strip()
    visible_name = original_name if original_name and (project / "sources" / original_name).exists() else text_name
    current_path = project_file_path(f"sources/{visible_name}")
    new_stem = Path(requested_filename(current_path, new_name)).stem
    if not new_stem:
        raise HTTPException(status_code=400, detail="请输入有效的文件名。")
    updated_original_name = f"{new_stem}{Path(original_name or visible_name).suffix}" if original_name else ""
    updated_text_name = f"{new_stem}.txt" if text_name else ""
    planned: dict[Path, Path] = {}
    if original_name:
        src = project / "sources" / original_name
        dst = project / "sources" / updated_original_name
        if src.exists():
            planned[src] = dst
    if text_name:
        src = project / "sources" / text_name
        dst = project / "sources" / updated_text_name
        if src.exists():
            planned[src] = dst
    for src, dst in planned.items():
        ensure_same_project_path(project, dst)
        if dst.exists() and dst != src:
            raise HTTPException(status_code=409, detail="目标文件名已存在。")
    for src, dst in planned.items():
        if src != dst:
            src.rename(dst)
    items = list_source_entries(project)
    for item in items:
        if item is entry or (
            item.get("filename") == entry.get("filename")
            and item.get("text_file") == entry.get("text_file")
        ):
            item["filename"] = updated_original_name or updated_text_name
            item["text_file"] = updated_text_name or updated_original_name
            break
    write_source_entries(project, items)
    final_name = updated_original_name or updated_text_name
    return {
        "filename": final_name,
        "relative_path": f"sources/{final_name}",
        "category": "Sources",
    }


def delete_source_managed_file(project: Path, entry: dict[str, Any]) -> None:
    original_name = str(entry.get("filename", "")).strip()
    text_name = str(entry.get("text_file", "")).strip()
    for name in {original_name, text_name}:
        if not name:
            continue
        path = project / "sources" / name
        if path.exists() and path.is_file():
            path.unlink()
    items = [
        item
        for item in list_source_entries(project)
        if not (
            item.get("filename") == entry.get("filename")
            and item.get("text_file") == entry.get("text_file")
        )
    ]
    write_source_entries(project, items)


def rename_project_file(project: Path, relative_path: str, new_name: str) -> dict[str, Any]:
    current = project_file_path(relative_path)
    category = project_file_category(project, current)
    actions = project_file_actions(project, current, category)
    if not actions["can_rename"]:
        raise HTTPException(status_code=409, detail="这个文件当前不允许重命名。")
    source_entry = source_entry_for_path(project, current) if category == "Sources" else None
    if source_entry:
        return rename_source_managed_file(project, source_entry, new_name)
    next_name = requested_filename(current, new_name)
    target = current.with_name(next_name)
    ensure_same_project_path(project, target)
    if target.exists() and target != current:
        raise HTTPException(status_code=409, detail="目标文件名已存在。")
    if target != current:
        current.rename(target)
    old_relative = project_relative_path(current, project)
    new_relative = project_relative_path(target, project)
    if category == "Manuscript" and current.suffix.lower() == ".qmd":
        update_active_manuscript_reference(project, old_relative, new_relative)
        migrate_editor_chat_history(project, old_relative, new_relative)
    return {
        "filename": target.name,
        "relative_path": new_relative,
        "category": project_file_category(project, target),
    }


def move_project_file(project: Path, relative_path: str, target_category: str) -> dict[str, Any]:
    current = project_file_path(relative_path)
    current_category = project_file_category(project, current)
    actions = project_file_actions(project, current, current_category)
    if target_category not in actions["move_targets"]:
        raise HTTPException(status_code=409, detail="这个文件当前不能拖到该分组。")
    target_dir = target_directory_for_category(project, target_category)
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / current.name
    ensure_same_project_path(project, target)
    if target.exists() and target != current:
        raise HTTPException(status_code=409, detail="目标位置已存在同名文件。")
    if target != current:
        current.rename(target)
    return {
        "filename": target.name,
        "relative_path": project_relative_path(target, project),
        "category": target_category,
    }


def delete_project_file(project: Path, relative_path: str) -> None:
    current = project_file_path(relative_path)
    category = project_file_category(project, current)
    actions = project_file_actions(project, current, category)
    if not actions["can_delete"]:
        if category == "Manuscript" and current.suffix.lower() == ".qmd" and len(manuscript_paths(project)) <= 1:
            raise HTTPException(status_code=409, detail="至少保留一个 qmd 文稿。")
        raise HTTPException(status_code=409, detail="这个文件当前不允许删除。")
    source_entry = source_entry_for_path(project, current) if category == "Sources" else None
    if source_entry:
        delete_source_managed_file(project, source_entry)
        return
    old_relative = project_relative_path(current, project)
    current.unlink()
    if category == "Manuscript" and current.suffix.lower() == ".qmd":
        update_active_manuscript_reference(project, old_relative, None)
        migrate_editor_chat_history(project, old_relative, None)


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def append_jsonl(path: Path, item: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(item, ensure_ascii=False) + "\n")


def read_jsonl(path: Path, limit: int = 20) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows = []
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        if not line.strip():
            continue
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return rows[-limit:]


def excerpt(text: str, limit: int = 700) -> str:
    compact = re.sub(r"\s+", " ", text).strip()
    if len(compact) <= limit:
        return compact
    return compact[:limit].rstrip() + "..."


def editor_asset_excerpt(path: Path, limit: int = 900) -> str:
    suffix = path.suffix.lower()
    if suffix not in {".md", ".txt", ".json", ".mmd", ".qmd"}:
        return ""
    try:
        raw = path.read_text(encoding="utf-8", errors="ignore").strip()
    except OSError:
        return ""
    if not raw:
        return ""
    if suffix == ".json":
        try:
            payload = json.loads(raw)
        except json.JSONDecodeError:
            return excerpt(raw, limit=limit)
        if isinstance(payload, dict):
            fields = [
                payload.get("prompt"),
                payload.get("summary"),
                payload.get("content"),
                payload.get("generated_code"),
            ]
            metadata = payload.get("metadata")
            if isinstance(metadata, dict):
                fields.extend([
                    metadata.get("analysis_title"),
                    metadata.get("summary"),
                    metadata.get("content"),
                    metadata.get("data_result"),
                    metadata.get("insert_paragraph"),
                    metadata.get("figure_relative_path"),
                ])
            raw = "\n".join(str(item).strip() for item in fields if str(item or "").strip())
            if not raw:
                raw = json.dumps(payload, ensure_ascii=False)
    return excerpt(raw, limit=limit)


def build_editor_asset_inventory(project: Path) -> dict[str, Any]:
    inventory = list_project_files(project)
    categories = {item.get("category", ""): item.get("files", []) for item in inventory}
    data_files = [
        {
            "name": item.get("name", ""),
            "relative_path": item.get("relative_path", ""),
        }
        for item in categories.get("Data", [])[:12]
    ]
    figures = [
        {
            "name": item.get("name", ""),
            "relative_path": item.get("relative_path", ""),
        }
        for item in categories.get("Figures", [])[:12]
    ]
    sources = [
        {
            "filename": item.get("filename", ""),
            "text_file": item.get("text_file", ""),
            "downloaded_original": bool(item.get("downloaded_original")),
        }
        for item in read_source_index()[:12]
    ]

    outputs = []
    output_paths = [
        path for path in sorted((project / "outputs").rglob("*"))
        if path.is_file() and path.name != ".gitkeep"
    ][:12]
    for path in output_paths:
        relative_path = str(path.relative_to(project))
        entry = {
            "name": path.name,
            "relative_path": relative_path,
            "extension": path.suffix.lower().lstrip(".") or "file",
        }
        if relative_path:
            asset_excerpt = editor_asset_excerpt(path)
            if asset_excerpt:
                entry["excerpt"] = asset_excerpt
        outputs.append(entry)

    return {
        "data_files": data_files,
        "figures": figures,
        "sources": sources,
        "outputs": outputs,
    }


def memory_summary_text(project: Path, limit: int = 4000) -> str:
    summary = project / "memory" / "summary.md"
    if not summary.exists():
        return ""
    text = summary.read_text(encoding="utf-8", errors="ignore").strip()
    if len(text) <= limit:
        return text
    return text[-limit:]


def memory_overview(project: Path) -> dict[str, Any]:
    conversations = read_jsonl(project / "memory" / "conversations.jsonl", limit=8)
    changes = read_jsonl(project / "memory" / "changes.jsonl", limit=8)
    return {
        "conversation_count": len(read_jsonl(project / "memory" / "conversations.jsonl", limit=100000)),
        "change_count": len(read_jsonl(project / "memory" / "changes.jsonl", limit=100000)),
        "recent_conversations": conversations,
        "recent_changes": changes,
        "summary": memory_summary_text(project, limit=1200),
    }


def append_memory_summary(project: Path, line: str) -> None:
    summary = project / "memory" / "summary.md"
    summary.parent.mkdir(parents=True, exist_ok=True)
    with summary.open("a", encoding="utf-8") as handle:
        handle.write(f"\n- {now_iso()} {line}\n")


def build_memory_context(project: Path) -> dict[str, Any]:
    conversations = read_jsonl(project / "memory" / "conversations.jsonl", limit=6)
    changes = read_jsonl(project / "memory" / "changes.jsonl", limit=6)
    return {
        "summary": memory_summary_text(project, limit=1800),
        "recent_ai_interactions": conversations,
        "recent_accepted_or_rejected_edits": changes,
    }


def log_ai_suggestion(
    project: Path,
    suggestion_id: str,
    req: SuggestRequest,
    suggestion: dict[str, Any],
    sources: list[dict[str, str]],
    raw_text: str,
) -> None:
    item = {
        "id": suggestion_id,
        "timestamp": now_iso(),
        "type": "suggestion",
        "instruction": req.instruction,
        "selected_text": excerpt(req.selected_text),
        "suggestion": suggestion,
        "source_files": sorted({source["filename"] for source in sources}),
        "raw_excerpt": excerpt(raw_text, limit=900),
        "status": "proposed",
    }
    append_jsonl(project / "memory" / "conversations.jsonl", item)
    append_memory_summary(
        project,
        f"AI suggested a revision for: \"{excerpt(req.selected_text, 160)}\"",
    )


def log_change(project: Path, item: dict[str, Any]) -> None:
    append_jsonl(project / "memory" / "changes.jsonl", item)
    status = item.get("status", "updated")
    if status == "accepted":
        append_memory_summary(project, f"Accepted edit: \"{excerpt(item.get('original_segment', ''), 120)}\" -> \"{excerpt(item.get('replacement', ''), 120)}\"")
    elif status == "rejected":
        append_memory_summary(project, f"Rejected suggestion for: \"{excerpt(item.get('original_segment', ''), 160)}\"")


def extract_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="需要安装 pypdf 才能读取 PDF。") from exc
    reader = PdfReader(str(path))
    return "\n\n".join(page.extract_text() or "" for page in reader.pages)


def extract_docx(path: Path) -> str:
    try:
        from docx import Document
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="需要安装 python-docx 才能读取 Word 文件。") from exc
    document = Document(str(path))
    return "\n".join(paragraph.text for paragraph in document.paragraphs if paragraph.text.strip())


def extract_csv(path: Path) -> str:
    rows = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        for idx, row in enumerate(reader):
            if idx >= 200:
                rows.append(["..."])
                break
            rows.append(row)
    return "\n".join("\t".join(cell for cell in row) for row in rows)


def extract_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="需要安装 openpyxl 才能读取 Excel 文件。") from exc
    workbook = load_workbook(str(path), read_only=True, data_only=True)
    parts = []
    for sheet in workbook.worksheets[:5]:
        parts.append(f"工作表: {sheet.title}")
        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if idx >= 100:
                parts.append("...")
                break
            parts.append("\t".join("" if value is None else str(value) for value in row))
    return "\n".join(parts)


def extract_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return extract_pdf(path)
    if suffix == ".docx":
        return extract_docx(path)
    if suffix == ".csv":
        return extract_csv(path)
    if suffix in [".xlsx", ".xlsm"]:
        return extract_xlsx(path)
    raise HTTPException(status_code=400, detail="目前支持 PDF、Word .docx、CSV 和 Excel .xlsx。")


def chunks(text: str, size: int = 1200) -> list[str]:
    clean = re.sub(r"\s+", " ", text).strip()
    return [clean[i : i + size] for i in range(0, len(clean), size) if clean[i : i + size].strip()]


def search_sources(query: str, limit: int = 5) -> list[dict[str, str]]:
    project = workspace_path()
    terms = [term.lower() for term in re.findall(r"[\w\-]{2,}", query)]
    scored = []
    for item in read_source_index():
        text_path = project / "sources" / item.get("text_file", "")
        if not text_path.exists():
            continue
        for chunk in chunks(text_path.read_text(encoding="utf-8", errors="ignore")):
            lower = chunk.lower()
            score = sum(lower.count(term) for term in terms)
            if not terms and query[:12] in chunk:
                score = 1
            if score > 0:
                scored.append((score, item["filename"], chunk))
    scored.sort(key=lambda row: row[0], reverse=True)
    return [{"filename": filename, "text": text} for _, filename, text in scored[:limit]]


def read_source_entry(project: Path, filename: str = "", text_file: str = "") -> dict[str, Any] | None:
    wanted_filename = str(filename or "").strip()
    wanted_text_file = str(text_file or "").strip()
    if not wanted_filename and not wanted_text_file:
        return None
    for item in read_source_index():
        if wanted_filename and item.get("filename") == wanted_filename:
            return item
        if wanted_text_file and item.get("text_file") == wanted_text_file:
            return item
    return None


def literature_candidate_from_source_entry(project: Path, source_entry: dict[str, Any] | None) -> dict[str, Any]:
    if not source_entry:
        return {}
    filename = str(source_entry.get("filename", "") or source_entry.get("text_file", "") or "").strip()
    text_file = str(source_entry.get("text_file", "")).strip()
    title = re.sub(r"[_-]+", " ", Path(filename).stem).strip() or filename or "Imported source"
    excerpt_text = ""
    if text_file:
        text_path = project / "sources" / text_file
        if text_path.exists():
            excerpt_text = excerpt(text_path.read_text(encoding="utf-8", errors="ignore"), limit=6000)
    return {
        "title": title,
        "authors": [],
        "year": "",
        "venue": "Project source",
        "abstract": "",
        "source_url": str(source_entry.get("source_url", "") or "").strip(),
        "download_url": "",
        "doi": "",
        "openalex_id": "",
        "excerpt": excerpt_text,
    }


def focused_source_hits(project: Path, source_entry: dict[str, Any] | None, query: str, limit: int = 5) -> list[dict[str, str]]:
    if not source_entry:
        return []
    text_file = str(source_entry.get("text_file", "")).strip()
    if not text_file:
        return []
    text_path = project / "sources" / text_file
    if not text_path.exists():
        return []

    terms = [term.lower() for term in re.findall(r"[\w\-]{2,}", query)]
    candidates = []
    for chunk in chunks(text_path.read_text(encoding="utf-8", errors="ignore"), size=1200):
        lower = chunk.lower()
        score = sum(lower.count(term) for term in terms) if terms else 0
        if not terms:
            score = 1
        if score > 0:
            candidates.append((score, chunk))
    candidates.sort(key=lambda row: row[0], reverse=True)
    selected = candidates[:limit] if candidates else []
    if not selected:
        selected = [(1, chunk) for chunk in chunks(text_path.read_text(encoding="utf-8", errors="ignore"), size=1200)[:limit]]
    return [{"filename": source_entry.get("filename", text_file), "text": text} for _, text in selected[:limit]]


def build_literature_chat_context(req: ChatRequest, project: Path) -> dict[str, Any]:
    document = paper_path().read_text(encoding="utf-8")
    outline = outline_from_document(document)
    cache_id = str(req.context.get("cache_id", "")).strip()
    source_filename = str(req.context.get("filename", "")).strip()
    text_file = str(req.context.get("text_file", "")).strip()
    should_search_fresh = not cache_id and not source_filename and not text_file
    bundle = search_literature_candidates(req.message) if should_search_fresh else {
        "candidate": {},
        "search_results": [],
        "scholar_search_url": "",
        "query_kind": "focus",
    }
    candidate = dict(bundle["candidate"] or {})
    cached_analysis: dict[str, Any] = {}

    if cache_id:
        try:
            cached = load_cached_literature(cache_id)
            candidate = dict(cached.get("candidate") or candidate)
            cached_analysis = dict(cached.get("analysis") or {})
        except HTTPException:
            pass

    source_entry = read_source_entry(project, filename=source_filename, text_file=text_file)
    default_candidate = literature_candidate_from_source_entry(project, source_entry)
    for key, value in default_candidate.items():
        if not candidate.get(key):
            candidate[key] = value
    source_hits = focused_source_hits(project, source_entry, req.message or candidate.get("title", ""), limit=4)
    if not source_hits and (req.message or candidate.get("title")):
        source_hits = search_sources(f"{candidate.get('title', '')}\n{req.message}".strip(), limit=4)

    source_focus = None
    if source_entry:
        source_focus = {
            "filename": source_entry.get("filename", ""),
            "text_file": source_entry.get("text_file", ""),
            "title": candidate.get("title", "") or source_entry.get("filename", ""),
            "downloaded_original": bool(source_entry.get("downloaded_original")),
        }

    candidate.setdefault("excerpt", "")
    if not bundle.get("scholar_search_url") and candidate.get("title"):
        bundle["scholar_search_url"] = build_google_scholar_search_url(candidate.get("title", ""))
    return {
        "candidate": candidate,
        "cached_analysis": cached_analysis,
        "search_results": bundle.get("search_results", []),
        "scholar_search_url": bundle.get("scholar_search_url", ""),
        "query_kind": bundle.get("query_kind", "query"),
        "outline": outline,
        "source_focus": source_focus,
        "imported_source_excerpts": source_hits,
    }


def new_memory_id(prefix: str) -> str:
    return f"{prefix}-{datetime.now().strftime('%Y%m%d%H%M%S%f')}"


VALID_CHAT_TOOLS = {"literature", "data", "mindmap", "brief", "editor"}


def normalize_chat_key(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", str(value or "").strip()).strip("._-")
    return cleaned[:80]


def chat_history_path(tool: str, project: Optional[Path] = None, chat_key: str = "") -> Path:
    safe_key = normalize_chat_key(chat_key)
    suffix = f"__{safe_key}" if safe_key else ""
    return (project or workspace_path()) / "memory" / "chats" / f"{tool}{suffix}.jsonl"


def load_chat_history(tool: str, project: Optional[Path] = None, chat_key: str = "") -> list[dict[str, Any]]:
    path = chat_history_path(tool, project, chat_key=chat_key)
    if not path.exists():
        return []
    rows = []
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return rows


def append_chat_messages(
    tool: str,
    project: Path,
    user_msg: dict[str, Any],
    assistant_msg: dict[str, Any],
    chat_key: str = "",
) -> None:
    path = chat_history_path(tool, project, chat_key=chat_key)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(user_msg, ensure_ascii=False) + "\n")
        handle.write(json.dumps(assistant_msg, ensure_ascii=False) + "\n")


def clear_chat_history(tool: str, project: Path, chat_key: str = "") -> None:
    path = chat_history_path(tool, project, chat_key=chat_key)
    if path.exists():
        path.write_text("", encoding="utf-8")


def contextual_document_excerpt(document: str, selected_text: str, radius: int = 1200) -> str:
    if not document:
        return ""
    needle = str(selected_text or "").strip()
    if not needle:
        return excerpt(document, limit=6000)
    start = document.find(needle)
    if start < 0:
        return excerpt(document, limit=6000)
    end = start + len(needle)
    snippet_start = max(0, start - radius)
    snippet_end = min(len(document), end + radius)
    snippet = document[snippet_start:snippet_end]
    if snippet_start > 0:
        snippet = f"...{snippet}"
    if snippet_end < len(document):
        snippet = f"{snippet}..."
    return snippet


def manuscript_section_snapshots(document: str, per_section_limit: int = 420, max_sections: int = 12) -> list[dict[str, str]]:
    snapshots: list[dict[str, str]] = []
    active_title = "Preamble"
    active_level = 0
    active_lines: list[str] = []

    def flush_section() -> None:
        nonlocal active_lines
        body = "\n".join(active_lines).strip()
        if not body and active_level > 0:
            snapshots.append({"title": active_title, "level": str(active_level), "excerpt": ""})
        elif body:
            snapshots.append(
                {
                    "title": active_title,
                    "level": str(active_level),
                    "excerpt": excerpt(body, limit=per_section_limit),
                }
            )
        active_lines = []

    for line in document.splitlines():
        match = re.match(r"^(#{1,6})\s+(.+)$", line.strip())
        if match:
            if active_lines or snapshots:
                flush_section()
            active_title = match.group(2).strip()
            active_level = len(match.group(1))
            continue
        active_lines.append(line)
    if active_lines or not snapshots:
        flush_section()
    return snapshots[:max_sections]


def build_editor_chat_context(req: ChatRequest, project: Path) -> tuple[dict[str, Any], list[dict[str, str]]]:
    document = str(req.context.get("document", "") or paper_path().read_text(encoding="utf-8"))
    selected_text = str(req.context.get("selected_text", "")).strip()
    source_hits = search_sources(f"{req.message}\n{selected_text}".strip()) if (req.message or selected_text) else []
    recent_turns = []
    for msg in req.history[-6:]:
        item: dict[str, Any] = {
            "role": msg.role,
            "content": msg.content,
        }
        selected_from_context = str(msg.context.get("selected_text", "")) if msg.context else ""
        selected_from_result = str(msg.result.get("selected_text", "")) if msg.result else ""
        effective_selected_text = selected_from_context.strip() or selected_from_result.strip()
        if effective_selected_text:
            item["selected_text"] = excerpt(effective_selected_text, limit=600)
        if msg.result:
            rewritten = str(msg.result.get("rewritten_text", "")).strip()
            rationale = str(msg.result.get("rationale", "")).strip()
            tool_results = msg.result.get("tool_results") if isinstance(msg.result.get("tool_results"), list) else []
            if rewritten:
                item["rewritten_text"] = excerpt(rewritten, limit=1200)
            if rationale:
                item["rationale"] = rationale
            if tool_results:
                item["tool_results"] = [
                    {
                        "type": result.get("type", ""),
                        "status": result.get("status", ""),
                        "summary": excerpt(
                            str(
                                result.get("summary")
                                or result.get("literature_review")
                                or result.get("data_result")
                                or result.get("title")
                                or result.get("error")
                                or ""
                            ),
                            limit=240,
                        ),
                    }
                    for result in tool_results[:3]
                ]
        recent_turns.append(item)
    return (
        {
            "active_manuscript": req.context.get("active_manuscript") or paper_path().name,
            "selection_mode": "explicit" if selected_text else "auto",
            "selected_text": selected_text,
            "manuscript_excerpt": contextual_document_excerpt(document, selected_text),
            "manuscript_section_snapshots": manuscript_section_snapshots(document),
            "project_asset_inventory": build_editor_asset_inventory(project),
            "paper_outline": outline_from_document(document),
            "local_sources": source_hits,
            "project_memory": build_memory_context(project),
            "recent_editor_turns": recent_turns,
        },
        source_hits,
    )


def upsert_source_index_entry(entry: dict[str, Any]) -> None:
    items = [item for item in read_source_index() if item.get("filename") != entry.get("filename")]
    items.append(entry)
    write_source_index(items)


def execute_editor_tool_actions(
    project: Path,
    provider: Any,
    settings: dict[str, Any],
    tool_actions: list[dict[str, Any]],
    user_message: str,
    document: str,
    outline: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []

    for action in tool_actions[:2]:
        action_type = str(action.get("type", "")).strip()
        base_result = {
            "type": action_type,
            "reason": str(action.get("reason", "")).strip(),
            "status": "ok",
        }
        try:
            if action_type == "import_literature":
                query = str(action.get("query", "")).strip()
                bundle = search_literature_candidates(query)
                candidate = bundle["candidate"]
                candidate.setdefault("excerpt", "")
                prompt = build_literature_prompt(query, candidate, outline)
                model_result = provider.analyze_literature(settings, prompt)
                analysis = model_result["analysis"]
                analysis["title"] = analysis.get("title") or candidate.get("title", "")
                analysis["authors"] = analysis.get("authors") or candidate.get("authors", [])
                analysis["year"] = analysis.get("year") or candidate.get("year", "")
                analysis["venue"] = analysis.get("venue") or candidate.get("venue", "")
                output_relative_path = save_literature_review_output(project, candidate, analysis)
                cached_payload = {
                    "query": query,
                    "candidate": candidate,
                    "analysis": analysis,
                    "raw": model_result.get("raw", ""),
                    "created_at": now_iso(),
                }
                cache_id = cache_literature_result(cached_payload)
                source_entry = import_literature_source(cache_id, project, bool(action.get("download_original")))
                upsert_source_index_entry(source_entry)
                results.append(
                    {
                        **base_result,
                        "query": query,
                        "cache_id": cache_id,
                        "candidate_title": candidate.get("title", ""),
                        "source_filename": source_entry.get("filename", ""),
                        "downloaded_original": bool(source_entry.get("downloaded_original")),
                        "literature_review": analysis.get("literature_review", ""),
                        "summary": analysis.get("summary", ""),
                        "output_relative_path": output_relative_path,
                    }
                )
                continue

            if action_type == "create_data_figure":
                relative_path = str(action.get("data_relative_path", "")).strip()
                prompt = str(action.get("prompt", "")).strip() or user_message
                analysis_result = run_data_analysis_skill(project, provider, settings, relative_path, prompt, outline)["analysis"]
                results.append(
                    {
                        **base_result,
                        "data_relative_path": relative_path,
                        "figure_relative_path": analysis_result.get("figure_relative_path", ""),
                        "figure_title": analysis_result.get("figure_title", ""),
                        "figure_caption": analysis_result.get("figure_caption", ""),
                        "suggested_section": analysis_result.get("suggested_section", ""),
                        "insert_paragraph": analysis_result.get("insert_paragraph", ""),
                        "data_result": analysis_result.get("data_result", ""),
                        "output_relative_path": analysis_result.get("record_relative_path", ""),
                    }
                )
                continue

            if action_type == "create_brief":
                brief_result = run_brief_skill(
                    project,
                    provider,
                    settings,
                    str(action.get("prompt", "")).strip() or user_message,
                    str(action.get("format", "summary")).strip() or "summary",
                    str(action.get("scope_heading", "")).strip(),
                    document,
                    outline,
                )["brief"]
                results.append(
                    {
                        **base_result,
                        "target_format": brief_result.get("target_format", ""),
                        "title": brief_result.get("title", ""),
                        "summary": brief_result.get("summary", ""),
                        "one_liner": brief_result.get("one_liner", ""),
                        "key_messages": brief_result.get("key_messages", []),
                        "output_relative_path": brief_result.get("output_relative_path", ""),
                    }
                )
                continue

            results.append(
                {
                    **base_result,
                    "status": "error",
                    "error": f"Unsupported tool action: {action_type or 'unknown'}",
                }
            )
        except HTTPException as exc:
            results.append(
                {
                    **base_result,
                    "status": "error",
                    "error": str(exc.detail),
                }
            )
        except Exception as exc:
            results.append(
                {
                    **base_result,
                    "status": "error",
                    "error": str(exc),
                }
            )

    return results


def build_chat_messages(
    history: list[dict[str, Any]],
    new_message: str,
    embedded_context: str,
) -> list[dict[str, str]]:
    recent = history[-16:] if len(history) > 16 else history
    messages: list[dict[str, str]] = []
    for msg in recent:
        messages.append({"role": msg["role"], "content": str(msg.get("content", ""))})
    user_content = new_message
    if embedded_context:
        user_content = f"{new_message}\n\nContext:\n{embedded_context}"
    messages.append({"role": "user", "content": user_content})
    return messages


def build_ai_prompt(req: SuggestRequest, document: str, project: Path) -> tuple[str, list[dict[str, str]]]:
    outline = outline_from_document(document)
    context_query = f"{req.instruction}\n{req.selected_text}"
    source_hits = search_sources(context_query)
    payload = {
        "user_instruction": req.instruction,
        "selected_text": req.selected_text,
        "paper_outline": outline,
        "local_sources": source_hits,
        "project_memory": build_memory_context(project),
    }
    return (
        json.dumps(payload, ensure_ascii=False, indent=2),
        source_hits,
    )


def build_ai_trace(settings: dict[str, Any], source_hits: list[dict[str, str]], project: Path) -> dict[str, Any]:
    memory_context = build_memory_context(project)
    return {
        "provider": settings.get("provider", "openai"),
        "model": settings.get("model", ""),
        "reasoning_effort": settings.get("reasoning", ""),
        "source_count": len(source_hits),
        "source_files": sorted({source["filename"] for source in source_hits}),
        "memory_summary_used": bool(memory_context.get("summary")),
        "recent_ai_interactions_used": len(memory_context.get("recent_ai_interactions", [])),
        "recent_edits_used": len(memory_context.get("recent_accepted_or_rejected_edits", [])),
    }


def _chat_embedded_context(tool: str, req: ChatRequest, project: Path) -> str:
    """Build the JSON context string embedded in the new user message."""
    document = paper_path().read_text(encoding="utf-8")
    outline = outline_from_document(document)

    if tool == "mindmap":
        return json.dumps(
            {"outline": outline, "document_excerpt": document[:4000]},
            ensure_ascii=False,
        )

    if tool == "brief":
        from .analysis_skills import extract_section_text
        format_val = req.context.get("format", "ppt")
        scope = req.context.get("scope_heading", "")
        scoped = extract_section_text(document, scope).strip() or document
        return json.dumps(
            {"target_format": format_val, "scope_heading": scope, "outline": outline, "source_excerpt": scoped[:6000]},
            ensure_ascii=False,
        )

    if tool == "data":
        relative_path = req.context.get("relative_path", "")
        if not relative_path:
            raise HTTPException(status_code=400, detail="data tool 需要提供 relative_path。")
        data_path = safe_project_file(project, relative_path, SUPPORTED_DATA_SUFFIXES)
        df = load_tabular_data(data_path)
        profile = dataframe_profile(df)
        previous_code = ""
        for msg in reversed(req.history):
            if msg.role == "assistant" and msg.result and msg.result.get("generated_code"):
                previous_code = msg.result["generated_code"]
                break
        ctx: dict[str, Any] = {
            "data_file": data_path.name,
            "manuscript_outline": outline,
            "dataset_profile": profile,
        }
        if previous_code:
            ctx["previous_code"] = previous_code
        return json.dumps(ctx, ensure_ascii=False)

    if tool == "literature":
        return json.dumps(build_literature_chat_context(req, project), ensure_ascii=False)

    if tool == "editor":
        payload, _source_hits = build_editor_chat_context(req, project)
        return json.dumps(payload, ensure_ascii=False)

    return ""


def run_chat_turn(
    tool: str,
    project: Path,
    provider: Any,
    settings: dict[str, Any],
    req: ChatRequest,
) -> dict[str, Any]:
    """Dispatch one chat turn to the appropriate skill and return the assistant message dict."""
    document = paper_path().read_text(encoding="utf-8")
    outline = outline_from_document(document)
    history_dicts = [msg.model_dump() for msg in req.history]
    embedded_context = _chat_embedded_context(tool, req, project)
    messages = build_chat_messages(history_dicts, req.message, embedded_context)

    if tool == "mindmap":
        result_data = chat_mindmap_turn(project, provider, settings, messages, outline, document)
        result = result_data["mindmap"]
        content = result.get("content") or result.get("summary", "已生成思维导图。")

    elif tool == "brief":
        target_format = req.context.get("format", "ppt")
        result_data = chat_brief_turn(project, provider, settings, messages, target_format, outline)
        result = result_data["brief"]
        content = result.get("content") or result.get("summary", "已生成展示摘要。")

    elif tool == "data":
        relative_path = req.context.get("relative_path", "")
        result_data = chat_data_turn(project, provider, settings, messages, relative_path, outline)
        result = result_data["analysis"]
        content = result.get("content") or result.get("summary", "已生成图表。")

    elif tool == "literature":
        from .providers import literature_instructions, parse_literature_json
        literature_context = build_literature_chat_context(req, project)
        candidate = literature_context["candidate"]
        cached_analysis = literature_context.get("cached_analysis") or {}
        raw_text = provider.generate_chat_json(settings, literature_instructions(settings), messages)
        analysis = parse_literature_json(raw_text)
        analysis["title"] = analysis.get("title") or candidate.get("title", "")
        analysis["authors"] = analysis.get("authors") or candidate.get("authors", [])
        analysis["year"] = analysis.get("year") or candidate.get("year", "")
        analysis["venue"] = analysis.get("venue") or candidate.get("venue", "")
        analysis["literature_review"] = analysis.get("literature_review") or cached_analysis.get("literature_review", "")
        output_relative_path = save_literature_review_output(project, candidate, analysis)
        cached_payload = {
            "query": req.message,
            "candidate": candidate,
            "analysis": analysis,
            "raw": raw_text,
            "created_at": now_iso(),
        }
        cache_id = cache_literature_result(cached_payload)
        content = analysis.get("content") or analysis.get("summary", "已完成文献分析。")
        result = {
            "analysis": analysis,
            "candidate": candidate,
            "search_results": literature_context.get("search_results", []),
            "scholar_search_url": literature_context.get("scholar_search_url", ""),
            "query_kind": literature_context.get("query_kind", "query"),
            "source_focus": literature_context.get("source_focus"),
            "cache_id": cache_id,
            "download_available": bool(candidate.get("download_url")),
            "output_relative_path": output_relative_path,
        }

    elif tool == "editor":
        from .providers import (
            editor_chat_instructions,
            editor_tool_planner_instructions,
            parse_editor_chat_json,
            parse_editor_tool_plan_json,
        )
        context_payload, source_hits = build_editor_chat_context(req, project)
        planner_messages = build_chat_messages(history_dicts, req.message, json.dumps(context_payload, ensure_ascii=False))
        planner_raw = provider.generate_chat_json(settings, editor_tool_planner_instructions(settings), planner_messages)
        tool_plan = parse_editor_tool_plan_json(planner_raw)
        tool_results = execute_editor_tool_actions(
            project,
            provider,
            settings,
            tool_plan.get("tool_actions", []),
            req.message,
            document,
            outline,
        )
        if tool_results:
            context_payload, source_hits = build_editor_chat_context(req, project)
        final_context = {
            **context_payload,
            "executed_tool_results": tool_results,
            "requested_tool_actions": tool_plan.get("tool_actions", []),
            "tool_planner_reason": tool_plan.get("reason", ""),
        }
        final_messages = build_chat_messages(history_dicts, req.message, json.dumps(final_context, ensure_ascii=False))
        raw_text = provider.generate_chat_json(settings, editor_chat_instructions(settings), final_messages)
        parsed = parse_editor_chat_json(raw_text)
        selected_text = str(req.context.get("selected_text", "")).strip() or str(parsed.get("selected_text", "")).strip()
        parsed["selected_text"] = selected_text
        parsed["tool_actions"] = tool_plan.get("tool_actions", [])
        parsed["tool_results"] = tool_results
        has_actionable_operations = bool(parsed.get("operations"))
        has_tool_side_effects = bool(tool_results)
        suggestion_id = new_memory_id("sug") if ((parsed.get("rewritten_text") and selected_text) or has_actionable_operations or has_tool_side_effects) else ""
        if suggestion_id:
            append_jsonl(
                project / "memory" / "conversations.jsonl",
                {
                    "id": suggestion_id,
                    "timestamp": now_iso(),
                    "type": "editor_chat",
                    "instruction": req.message,
                    "selected_text": excerpt(selected_text),
                    "suggestion": {
                        "selected_text": selected_text,
                        "rewritten_text": parsed.get("rewritten_text", ""),
                        "operations": parsed.get("operations", []),
                        "tool_actions": parsed.get("tool_actions", []),
                        "tool_results": parsed.get("tool_results", []),
                        "rationale": parsed.get("rationale", ""),
                        "process_summary": parsed.get("process_summary", []),
                        "risks": parsed.get("risks", []),
                        "citation_or_data_notes": parsed.get("citation_or_data_notes", []),
                        "confidence": parsed.get("confidence", "medium"),
                    },
                    "source_files": sorted({source["filename"] for source in source_hits}),
                    "raw_excerpt": excerpt(raw_text, limit=900),
                    "status": "proposed",
                    "chat_mode": True,
                    "active_manuscript": context_payload.get("active_manuscript", ""),
                },
            )
            append_memory_summary(
                project,
                f"Editor chat proposed an edit plan for: \"{excerpt(selected_text or summarize_editor_operations(parsed.get('operations', [])) or summarize_editor_tool_actions(parsed.get('tool_actions', [])), 160)}\"",
            )
        result = {
            **parsed,
            "selected_text": selected_text,
            "suggestion_id": suggestion_id,
            "trace": build_ai_trace(settings, source_hits, project),
        }
        content = parsed.get("content") or parsed.get("rationale") or "已完成这轮写作协作。"

    else:
        raise HTTPException(status_code=400, detail=f"未知 tool：{tool}")

    return {
        "id": new_memory_id("msg"),
        "role": "assistant",
        "timestamp": now_iso(),
        "content": content,
        "result": result,
    }


def configured_provider(settings: Optional[dict[str, Any]] = None):
    resolved_settings = settings or read_settings()
    provider = get_provider(resolved_settings.get("provider", "openai"), load_env_file())
    if not provider.configured:
        raise HTTPException(status_code=400, detail=f"请先在设置中填写 {provider.display_name} API Key。")
    return provider, resolved_settings


def prepare_ai_request(req: SuggestRequest) -> tuple[Any, dict[str, Any], str, Path, list[dict[str, str]]]:
    if not req.selected_text.strip():
        raise HTTPException(status_code=400, detail="请先在正文中选中一段要修改的文字。")
    provider, settings = configured_provider()
    project = workspace_path()
    document = req.document if req.document is not None else paper_path().read_text(encoding="utf-8")
    prompt, source_hits = build_ai_prompt(req, document, project)
    return provider, settings, prompt, project, source_hits




@app.on_event("startup")
def startup() -> None:
    ensure_settings()


@app.get("/api/project")
def get_project() -> dict[str, Any]:
    settings = read_settings()
    quarto_path = shutil.which("quarto")
    payload = {
        "workspace_configured": False,
        "workspace_error": "",
        "workspace_suggestion": str(DEFAULT_PROJECTS_ROOT.resolve()),
        "projects_root": "",
        "workspace": "",
        "active_project": "",
        "active_manuscript": "",
        "manuscripts": [],
        "projects": [],
        "paper_exists": False,
        "outline": [],
        "sources": [],
        "files": [],
        "memory": {
            "conversation_count": 0,
            "change_count": 0,
            "recent_conversations": [],
            "recent_changes": [],
        },
        "settings": config_settings_payload(),
        "quarto_available": quarto_path is not None,
        "quarto_path": quarto_path,
        "quarto_message": (
            "Quarto is installed and ready for Word export."
            if quarto_path
            else "Quarto was not found. Editing and AI still work; install Quarto to export Word/PDF."
        ),
    }
    configured_root = configured_projects_root()
    if configured_root is None:
        return payload
    payload["projects_root"] = str(configured_root)
    if configured_root.exists() and not configured_root.is_dir():
        payload["workspace_error"] = WORKSPACE_INVALID_MESSAGE
        return payload
    if not configured_root.exists():
        payload["workspace_error"] = WORKSPACE_MISSING_MESSAGE
        return payload
    ensure_workspace()
    project_id = active_project_id()
    project = workspace_path(project_id)
    active_manuscript = active_manuscript_relative_path(project_id)
    content = (project / active_manuscript).read_text(encoding="utf-8")
    payload.update(
        {
            "workspace_configured": True,
            "workspace": str(project),
            "active_project": project_id,
            "active_manuscript": active_manuscript,
            "manuscripts": manuscript_entries(project_id),
            "projects": existing_projects(),
            "paper_exists": bool(manuscript_paths(project)),
            "outline": outline_from_document(content),
            "sources": read_source_index(project_id),
            "files": list_project_files(project),
            "memory": memory_overview(project),
        }
    )
    return payload


@app.post("/api/project/create")
def create_project(req: ProjectCreate) -> dict[str, Any]:
    name = (req.name or "Thesis Draft").strip()
    project_id = slugify(name)
    parent = workspace_path().parent
    suffix = 2
    candidate = project_id
    target = (parent / candidate).resolve()
    while target.exists():
        candidate = f"{project_id}-{suffix}"
        target = (parent / candidate).resolve()
        suffix += 1
    ensure_project(title=name, seed_legacy=False, root=target)
    set_projects_root(str(target))
    return get_project()


@app.post("/api/project/open")
def open_project(req: ProjectOpen) -> dict[str, Any]:
    set_active_project(req.project_id)
    return get_project()


@app.post("/api/workspace/root")
def update_workspace_root(req: WorkspaceRootUpdate) -> dict[str, Any]:
    set_projects_root(req.path)
    return get_project()


@app.post("/api/workspace/root/choose")
def choose_workspace_root() -> dict[str, Any]:
    selected = choose_projects_root_dialog()
    if not selected:
        return {"ok": False, "cancelled": True}
    set_projects_root(str(selected))
    payload = get_project()
    payload["ok"] = True
    payload["cancelled"] = False
    return payload


@app.get("/api/document")
def get_document() -> dict[str, Any]:
    ensure_workspace()
    current = paper_path()
    return {
        "content": current.read_text(encoding="utf-8"),
        "relative_path": str(current.relative_to(workspace_path())),
        "filename": current.name,
    }


@app.post("/api/document/open")
def open_document(req: DocumentOpen) -> dict[str, Any]:
    ensure_workspace()
    relative_path = set_active_manuscript(req.relative_path)
    current = workspace_path() / relative_path
    return {
        "ok": True,
        "content": current.read_text(encoding="utf-8"),
        "relative_path": relative_path,
        "filename": current.name,
        "outline": outline_from_document(current.read_text(encoding="utf-8")),
    }


@app.post("/api/document/create")
def create_document(req: DocumentCreate) -> dict[str, Any]:
    ensure_workspace()
    current = create_manuscript_file(req.filename)
    content = current.read_text(encoding="utf-8")
    return {
        "ok": True,
        "content": content,
        "relative_path": str(current.relative_to(workspace_path())),
        "filename": current.name,
        "outline": outline_from_document(content),
    }


@app.get("/api/project/file")
def get_project_file(relative_path: str) -> FileResponse:
    ensure_workspace()
    return FileResponse(project_file_path(relative_path))


@app.post("/api/project/file/rename")
def rename_project_file_endpoint(req: ProjectFileRename) -> dict[str, Any]:
    ensure_workspace()
    project = workspace_path()
    result = rename_project_file(project, req.relative_path, req.new_name)
    return {"ok": True, **result}


@app.post("/api/project/file/move")
def move_project_file_endpoint(req: ProjectFileMove) -> dict[str, Any]:
    ensure_workspace()
    project = workspace_path()
    result = move_project_file(project, req.relative_path, req.target_category)
    return {"ok": True, **result}


@app.post("/api/project/file/delete")
def delete_project_file_endpoint(req: ProjectFileDelete) -> dict[str, Any]:
    ensure_workspace()
    project = workspace_path()
    delete_project_file(project, req.relative_path)
    return {"ok": True}


@app.post("/api/document")
def update_document(update: DocumentUpdate) -> dict[str, Any]:
    ensure_workspace()
    current = paper_path()
    current.write_text(update.content, encoding="utf-8")
    return {"ok": True, "relative_path": str(current.relative_to(workspace_path())), "outline": outline_from_document(update.content)}


@app.get("/api/settings")
def get_settings() -> dict[str, Any]:
    return config_settings_payload()


@app.get("/api/providers")
def get_providers() -> dict[str, Any]:
    return provider_payload(read_settings(), load_env_file())


@app.post("/api/settings")
def update_settings(update: SettingsUpdate) -> dict[str, Any]:
    settings = read_settings()
    for field in ["provider", "model", "reasoning", "reference_doc", "export_dir"]:
        value = getattr(update, field)
        if value:
            settings[field] = value
    if update.instruction is not None:
        settings["instruction"] = update.instruction.strip() or DEFAULT_AI_INSTRUCTION
    if update.provider and not update.model:
        defaults = DEFAULT_PROVIDER_MODELS.get(update.provider, [])
        if defaults:
            settings["model"] = defaults[0]["id"]
    env_updates: dict[str, str | None] = {}
    if update.api_key:
        env_updates["OPENAI_API_KEY"] = update.api_key
    if update.openai_api_key:
        env_updates["OPENAI_API_KEY"] = update.openai_api_key
    if update.deepseek_api_key:
        env_updates["DEEPSEEK_API_KEY"] = update.deepseek_api_key
    if update.deepseek_base_url:
        env_updates["DEEPSEEK_BASE_URL"] = update.deepseek_base_url
    if env_updates:
        update_env_values(env_updates)
    save_settings(settings)
    return get_settings()


@app.get("/api/memory")
def get_memory() -> dict[str, Any]:
    ensure_workspace()
    project = workspace_path()
    return memory_overview(project)


@app.post("/api/sources/import")
async def import_source(file: UploadFile = File(...)) -> dict[str, Any]:
    ensure_workspace()
    if not file.filename:
        raise HTTPException(status_code=400, detail="没有收到文件名。")
    project = workspace_path()
    safe_name = Path(file.filename).name
    target = project / "sources" / safe_name
    with target.open("wb") as handle:
        handle.write(await file.read())
    extracted = extract_text(target)
    if not extracted.strip():
        raise HTTPException(status_code=400, detail="文件已保存，但没有提取到可读取的文字。")
    text_file = f"{target.stem}.txt"
    (project / "sources" / text_file).write_text(extracted, encoding="utf-8")
    items = [item for item in read_source_index() if item.get("filename") != safe_name]
    entry = {
        "filename": safe_name,
        "text_file": text_file,
        "characters": len(extracted),
        "imported_at": datetime.now().isoformat(timespec="seconds"),
    }
    items.append(entry)
    write_source_index(items)
    return {"ok": True, "source": entry}


@app.post("/api/literature/analyze")
def analyze_literature(req: LiteratureAnalyzeRequest) -> dict[str, Any]:
    ensure_workspace()
    provider, settings = configured_provider()
    project = workspace_path()
    outline = outline_from_document(paper_path().read_text(encoding="utf-8"))
    query = str(req.query or "").strip()
    cache_id = str(req.cache_id or "").strip()
    source_filename = str(req.source_filename or "").strip()
    text_file = str(req.text_file or "").strip()
    source_entry = read_source_entry(project, filename=source_filename, text_file=text_file)
    source_focus = None
    cached_analysis: dict[str, Any] = {}

    if cache_id or source_entry:
        bundle = {
            "candidate": {},
            "search_results": [],
            "scholar_search_url": "",
            "query_kind": "focus",
        }
        candidate = {}
        if cache_id:
            try:
                cached = load_cached_literature(cache_id)
                candidate = dict(cached.get("candidate") or {})
                cached_analysis = dict(cached.get("analysis") or {})
            except HTTPException:
                pass
        default_candidate = literature_candidate_from_source_entry(project, source_entry)
        for key, value in default_candidate.items():
            if not candidate.get(key):
                candidate[key] = value
        if not candidate.get("title") and not query:
            raise HTTPException(status_code=400, detail="请选择项目资料，或输入论文标题、DOI、URL。")
        source_hits = focused_source_hits(project, source_entry, query or candidate.get("title", ""), limit=4)
        source_focus = (
            {
                "filename": source_entry.get("filename", ""),
                "text_file": source_entry.get("text_file", ""),
                "title": candidate.get("title", "") or source_entry.get("filename", ""),
                "downloaded_original": bool(source_entry.get("downloaded_original")),
            }
            if source_entry
            else None
        )
        if candidate.get("title"):
            bundle["scholar_search_url"] = build_google_scholar_search_url(candidate.get("title", ""))
        prompt = build_literature_prompt(
            query or candidate.get("title", ""),
            candidate,
            outline,
            imported_source_excerpts=source_hits,
            source_focus=source_focus,
            query_kind="focus",
        )
    else:
        if not query:
            raise HTTPException(status_code=400, detail="请输入论文标题、DOI、URL，或选择项目资料。")
        bundle = search_literature_candidates(query)
        candidate = bundle["candidate"]
        candidate.setdefault("excerpt", "")
        prompt = build_literature_prompt(query, candidate, outline, query_kind=bundle.get("query_kind", "query"))

    try:
        result = provider.analyze_literature(settings, prompt)
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"{provider.display_name} 调用失败：{exc}") from exc
    analysis = result["analysis"]
    analysis["title"] = analysis.get("title") or candidate.get("title", "")
    analysis["authors"] = analysis.get("authors") or candidate.get("authors", [])
    analysis["year"] = analysis.get("year") or candidate.get("year", "")
    analysis["venue"] = analysis.get("venue") or candidate.get("venue", "")
    analysis["literature_review"] = analysis.get("literature_review") or cached_analysis.get("literature_review", "")
    output_relative_path = save_literature_review_output(project, candidate, analysis)
    cached_payload = {
        "query": query or candidate.get("title", ""),
        "candidate": candidate,
        "analysis": analysis,
        "raw": result.get("raw", ""),
        "created_at": now_iso(),
    }
    cache_id = cache_literature_result(cached_payload)
    return {
        "ok": True,
        "cache_id": cache_id,
        "candidate": candidate,
        "search_results": bundle.get("search_results", []),
        "scholar_search_url": bundle.get("scholar_search_url", ""),
        "query_kind": bundle.get("query_kind", "query"),
        "source_focus": source_focus,
        "analysis": analysis,
        "download_available": bool(candidate.get("download_url")),
        "output_relative_path": output_relative_path,
    }


@app.post("/api/literature/import")
def import_literature(req: LiteratureImportRequest) -> dict[str, Any]:
    ensure_workspace()
    project = workspace_path()
    entry = import_literature_source(req.cache_id, project, req.download_original)
    items = [item for item in read_source_index() if item.get("filename") != entry["filename"]]
    items.append(entry)
    write_source_index(items)
    return {"ok": True, "source": entry}


@app.post("/api/analysis/data")
def analyze_data(req: DataAnalysisRequest) -> dict[str, Any]:
    ensure_workspace()
    provider, settings = configured_provider()
    project = workspace_path()
    document = paper_path().read_text(encoding="utf-8")
    outline = outline_from_document(document)
    try:
        result = run_data_analysis_skill(project, provider, settings, req.relative_path, req.prompt, outline)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"{provider.display_name} 数据分析失败：{exc}") from exc
    return {"ok": True, **result}


@app.post("/api/analysis/data/insert")
def insert_data_figure(req: DataFigureInsertRequest) -> dict[str, Any]:
    ensure_workspace()
    project = workspace_path()
    safe_figure = project_file_path(req.figure_relative_path)
    if safe_figure.suffix.lower() not in [".png", ".jpg", ".jpeg", ".svg", ".webp"]:
        raise HTTPException(status_code=400, detail="只能插入图片文件到文稿中。")
    current_document = paper_path()
    updated = insert_figure_into_manuscript(
        current_document.read_text(encoding="utf-8"),
        {
            "figure_relative_path": req.figure_relative_path,
            "figure_title": req.figure_title,
            "figure_caption": req.figure_caption,
            "figure_alt_text": req.figure_alt_text or "",
            "section_title": req.section_title or "",
            "introduction": req.introduction or "",
        },
    )
    current_document.write_text(updated, encoding="utf-8")
    return {
        "ok": True,
        "content": updated,
        "relative_path": str(current_document.relative_to(project)),
        "outline": outline_from_document(updated),
    }


@app.post("/api/analysis/mindmap")
def create_mindmap(req: MindmapRequest) -> dict[str, Any]:
    ensure_workspace()
    provider, settings = configured_provider()
    project = workspace_path()
    document = paper_path().read_text(encoding="utf-8")
    outline = outline_from_document(document)
    try:
        result = run_mindmap_skill(project, provider, settings, req.prompt, outline, document)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"{provider.display_name} 思维导图生成失败：{exc}") from exc
    return {"ok": True, **result}


@app.post("/api/analysis/mindmap/insert")
def insert_mindmap(req: MindmapInsertRequest) -> dict[str, Any]:
    ensure_workspace()
    project = workspace_path()
    current_document = paper_path()
    content = current_document.read_text(encoding="utf-8")
    updated = insert_mermaid_into_manuscript(content, req.quarto_block, req.section_title or "")
    current_document.write_text(updated, encoding="utf-8")
    return {
        "ok": True,
        "content": updated,
        "relative_path": str(current_document.relative_to(project)),
        "outline": outline_from_document(updated),
    }


@app.post("/api/analysis/brief")
def create_brief(req: BriefRequest) -> dict[str, Any]:
    ensure_workspace()
    provider, settings = configured_provider()
    project = workspace_path()
    document = paper_path().read_text(encoding="utf-8")
    outline = outline_from_document(document)
    try:
        result = run_brief_skill(project, provider, settings, req.prompt, req.format, req.scope_heading, document, outline)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"{provider.display_name} 展示摘要生成失败：{exc}") from exc
    return {"ok": True, **result}


@app.get("/api/chat/{tool}")
def get_chat_history(tool: str, chat_key: str = "") -> dict[str, Any]:
    if tool not in VALID_CHAT_TOOLS:
        raise HTTPException(status_code=400, detail=f"未知 tool：{tool}")
    ensure_workspace()
    project = workspace_path()
    return {"ok": True, "tool": tool, "history": load_chat_history(tool, project, chat_key=chat_key)}


@app.post("/api/chat/{tool}")
def post_chat_turn(tool: str, req: ChatRequest) -> dict[str, Any]:
    if tool not in VALID_CHAT_TOOLS:
        raise HTTPException(status_code=400, detail=f"未知 tool：{tool}")
    ensure_workspace()
    provider, settings = configured_provider()
    project = workspace_path()
    chat_key = str(req.context.get("chat_key", "") or "")

    user_msg: dict[str, Any] = {
        "id": new_memory_id("msg"),
        "role": "user",
        "timestamp": now_iso(),
        "content": req.message,
        "context": req.context or {},
    }

    try:
        assistant_msg = run_chat_turn(tool, project, provider, settings, req)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"AI 调用失败：{exc}") from exc

    append_chat_messages(tool, project, user_msg, assistant_msg, chat_key=chat_key)
    return {"ok": True, "message": assistant_msg}


@app.delete("/api/chat/{tool}")
def delete_chat_history(tool: str, chat_key: str = "") -> dict[str, Any]:
    if tool not in VALID_CHAT_TOOLS:
        raise HTTPException(status_code=400, detail=f"未知 tool：{tool}")
    ensure_workspace()
    clear_chat_history(tool, workspace_path(), chat_key=chat_key)
    return {"ok": True}


@app.post("/api/ai/suggest")
def suggest(req: SuggestRequest) -> dict[str, Any]:
    provider, settings, prompt, project, source_hits = prepare_ai_request(req)
    suggestion_id = new_memory_id("sug")
    try:
        result = provider.create_suggestion(settings, prompt)
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"{provider.display_name} 调用失败：{exc}") from exc
    raw_text = result["raw"]
    parsed = result["suggestion"]
    trace = build_ai_trace(settings, source_hits, project)
    log_ai_suggestion(project, suggestion_id, req, parsed, source_hits, raw_text)
    return {"ok": True, "suggestion_id": suggestion_id, "suggestion": parsed, "raw": raw_text, "trace": trace}


@app.post("/api/ai/suggest/stream")
def suggest_stream(req: SuggestRequest) -> StreamingResponse:
    provider, settings, prompt, project, source_hits = prepare_ai_request(req)
    suggestion_id = new_memory_id("sug")

    def event_stream():
        try:
            for event in provider.stream_suggestion(settings, prompt):
                if event.get("type") == "delta":
                    yield f"data: {json.dumps({'type': 'delta', 'text': event.get('text', '')}, ensure_ascii=False)}\n\n"
                elif event.get("type") == "final":
                    raw_text = event.get("raw", "")
                    parsed = event.get("suggestion", {})
                    trace = build_ai_trace(settings, source_hits, project)
                    log_ai_suggestion(project, suggestion_id, req, parsed, source_hits, raw_text)
                    yield f"data: {json.dumps({'type': 'final', 'suggestion_id': suggestion_id, 'suggestion': parsed, 'raw': raw_text, 'trace': trace}, ensure_ascii=False)}\n\n"
        except Exception as exc:
            payload = {"type": "error", "message": f"{provider.display_name} 调用失败：{exc}"}
            yield f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"

    return StreamingResponse(event_stream(), media_type="text/event-stream")


def summarize_editor_operations(operations: list[dict[str, Any]]) -> str:
    labels: list[str] = []
    for op in operations:
        op_type = str(op.get("type", "")).strip()
        if op_type == "replace_text":
            labels.append(f"Replace: {excerpt(str(op.get('target_text', '')), 90)}")
        elif op_type == "insert_under_heading":
            heading = str(op.get("section_title", "")).strip() or "document end"
            labels.append(f"Insert under {heading}: {excerpt(str(op.get('content', '')), 90)}")
        elif op_type == "insert_figure":
            heading = str(op.get("section_title", "")).strip() or "document end"
            labels.append(f"Insert figure {op.get('figure_relative_path', '')} under {heading}")
    return " | ".join(labels[:4]).strip()


def summarize_editor_tool_actions(actions: list[dict[str, Any]]) -> str:
    labels: list[str] = []
    for action in actions:
        action_type = str(action.get("type", "")).strip()
        if action_type == "import_literature":
            labels.append(f"Import literature: {action.get('query', '')}")
        elif action_type == "create_data_figure":
            labels.append(f"Create figure: {action.get('data_relative_path', '')}")
        elif action_type == "create_brief":
            labels.append(f"Create brief: {action.get('format', '')}")
    return " | ".join(labels[:4]).strip()


def apply_editor_operations(content: str, operations: list[dict[str, Any]], project: Path) -> str:
    updated = content
    for index, op in enumerate(operations, start=1):
        op_type = str(op.get("type", "")).strip()

        if op_type == "replace_text":
            target_text = str(op.get("target_text", "")).strip()
            replacement = str(op.get("replacement", "")).strip()
            if not target_text or not replacement:
                raise HTTPException(status_code=400, detail=f"第 {index} 个编辑动作缺少 target_text 或 replacement。")
            if target_text not in updated:
                raise HTTPException(status_code=409, detail=f"第 {index} 个替换目标已变化，请刷新文稿后重新生成。")
            updated = updated.replace(target_text, replacement, 1)
            continue

        if op_type == "insert_under_heading":
            insert_content = str(op.get("content", "")).strip()
            if not insert_content:
                raise HTTPException(status_code=400, detail=f"第 {index} 个插入动作缺少 content。")
            section_title = str(op.get("section_title", "")).strip()
            if section_title:
                updated = insert_block_into_section(updated, section_title, insert_content)
            else:
                updated = updated.rstrip() + "\n\n" + insert_content + "\n"
            continue

        if op_type == "insert_figure":
            figure_relative_path = str(op.get("figure_relative_path", "")).strip()
            if not figure_relative_path:
                raise HTTPException(status_code=400, detail=f"第 {index} 个 figure 动作缺少 figure_relative_path。")
            safe_figure = project_file_path(figure_relative_path)
            if safe_figure.suffix.lower() not in [".png", ".jpg", ".jpeg", ".svg", ".webp"]:
                raise HTTPException(status_code=400, detail="只能把图片文件作为 figure 插入文稿。")
            figure_title = str(op.get("figure_title", "")).strip() or safe_figure.stem.replace("-", " ").replace("_", " ").title()
            figure_caption = str(op.get("figure_caption", "")).strip() or figure_title
            updated = insert_figure_into_manuscript(
                updated,
                {
                    "figure_relative_path": figure_relative_path,
                    "figure_title": figure_title,
                    "figure_caption": figure_caption,
                    "figure_alt_text": str(op.get("figure_alt_text", "")).strip(),
                    "section_title": str(op.get("section_title", "")).strip(),
                    "introduction": str(op.get("introduction", "")).strip(),
                },
            )
            continue

        raise HTTPException(status_code=400, detail=f"不支持的编辑动作类型：{op_type or 'unknown'}。")

    return updated


@app.post("/api/ai/apply")
def apply_suggestion(req: ApplyRequest) -> dict[str, Any]:
    ensure_workspace()
    if not req.original_segment and not req.operations:
        raise HTTPException(status_code=400, detail="缺少可应用的编辑动作。")
    current_paper = paper_path()
    content = current_paper.read_text(encoding="utf-8")
    project = workspace_path()

    if req.operations:
        updated = apply_editor_operations(content, req.operations, project)
    else:
        original_segment = str(req.original_segment or "").strip()
        if not original_segment:
            raise HTTPException(status_code=400, detail="缺少原文片段，无法安全替换。")
        if original_segment not in content:
            raise HTTPException(status_code=409, detail="原文片段已经变化，请重新选择段落后再应用。")
        updated = content.replace(original_segment, req.replacement, 1)

    current_paper.write_text(updated, encoding="utf-8")
    log_change(
        project,
        {
            "id": new_memory_id("chg"),
            "suggestion_id": req.suggestion_id,
            "timestamp": now_iso(),
            "type": "edit",
            "status": "accepted",
            "original_segment": excerpt(str(req.original_segment or summarize_editor_operations(req.operations)), limit=900),
            "replacement": excerpt(str(req.replacement or summarize_editor_operations(req.operations)), limit=900),
            "operations": req.operations,
        },
    )
    return {"ok": True, "content": updated, "outline": outline_from_document(updated)}


@app.post("/api/ai/reject")
def reject_suggestion(req: RejectRequest) -> dict[str, Any]:
    ensure_workspace()
    original_segment = str(req.original_segment or "")
    operations = []
    if isinstance(req.suggestion, dict):
        maybe_ops = req.suggestion.get("operations")
        if isinstance(maybe_ops, list):
            operations = maybe_ops
    log_change(
        workspace_path(),
        {
            "id": new_memory_id("chg"),
            "suggestion_id": req.suggestion_id,
            "timestamp": now_iso(),
            "type": "edit",
            "status": "rejected",
            "original_segment": excerpt(original_segment or summarize_editor_operations(operations), limit=900),
            "suggestion": req.suggestion or {},
            "operations": operations,
        },
    )
    return {"ok": True, "memory": memory_overview(workspace_path())}


@app.post("/api/export/docx")
def export_docx() -> dict[str, Any]:
    ensure_workspace()
    if shutil.which("quarto") is None:
        raise HTTPException(status_code=400, detail="没有找到 Quarto。请先安装 Quarto 后再导出 Word。")
    settings = read_settings()
    project = workspace_path()
    current_manuscript = paper_path()
    export_dir = project / settings.get("export_dir", "outputs")
    export_dir.mkdir(parents=True, exist_ok=True)
    output_name = f"{current_manuscript.stem}-{datetime.now().strftime('%Y%m%d-%H%M%S')}.docx"
    command = ["quarto", "render", current_manuscript.name, "--to", "docx", "--output", output_name]
    reference_doc = project / settings.get("reference_doc", "templates/reference.docx")
    if reference_doc.exists():
        command.extend(["-M", f"reference-doc:{reference_doc}"])
    result = subprocess.run(command, cwd=project, capture_output=True, text=True, timeout=120)
    if result.returncode != 0:
        raise HTTPException(status_code=500, detail=f"Word 导出失败：{result.stderr or result.stdout}")
    generated = project / output_name
    final_path = export_dir / output_name
    if generated.exists() and generated != final_path:
        generated.replace(final_path)
    return {"ok": True, "path": str(final_path), "filename": output_name}
